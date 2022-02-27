import openpyxl

# ブックを取得
book = openpyxl.load_workbook('test.xlsx')
# シートを取得
sheet = book['Sheet1']

# 背景色を変更
fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='FFFF00', bgColor='FFFF00')
# シートに設定
for i in range(2,300):
    val = sheet.cell(row = i, column = 1).value
    if 'ta' in val:
        sheet[i][1].fill = fill

# 保存する
book.save('test.xlsx')