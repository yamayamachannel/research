from pathlib import Path
import math
import openpyxl
from matplotlib import pyplot

jp_font = "Yu Gothic" # 日本語フォント
pyplot.style.use("seaborn") # グラフスタイル

wb = openpyxl.load_workbook('naist2021pro.xlsx')
ws = wb['Sheet1']

def get_student_chat():
    lst_student_chat = []
    for i in range(2,84):
        sumchat = ws.cell(row = i, column = 3).value
        TAchat = ws.cell(row = i, column = 4).value
        lst_student_chat.append(sumchat - TAchat)
    return lst_student_chat
    # print(lst_student_chat)

def draw_hist_sales(sales_list, bin_edges):
    pyplot.hist(sales_list, bins=bin_edges, edgecolor="black", linewidth=1.0)
    pyplot.xlabel("受講生発言数", fontname = jp_font)
    pyplot.ylabel("度数", fontname = jp_font)
    
data_max, data_min, data_n = max(get_student_chat()), min(get_student_chat()), len(get_student_chat())
print("最大値：", data_max)
print("最小値：", data_min)
print("個数：", data_n)
print()

print("スタージェス：", 1 + math.log2(data_n))
print()

bin_min = int(input("階級の下限値＝"))
bin_max = int(input("階級の上限値＝"))
bin_w = int(input("階級の幅＝"))

# ヒストグラム作成
edges = range(bin_min, bin_max + bin_w, bin_w)
draw_hist_sales(get_student_chat(), edges)

# ヒストグラム出力
pyplot.show()
   