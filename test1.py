# チャット数とかを数えてexcelファイルに書き込む

# from collections import UserDict
# from pathlib import Path
# import math
import openpyxl
from matplotlib import pyplot
# from slacker import Chat

jp_font = "Yu Gothic" # 日本語フォント
pyplot.style.use("seaborn") # グラフスタイル

wb = openpyxl.load_workbook('test.xlsx')
ws1 = wb['Sheet1']
ws2 = wb['Sheet2']

def get_student_chat():
    userID = []
    chat = []
    for i in range(2,20):
        data1 = ws1.cell(row = i, column = 1).value
        data2 = ws1.cell(row = i, column = 2).value
        userID.append(data1)
        chat.append(data2)
    d = dict(zip(userID,chat))
    
    for j in range(1,7):
        chat1 = ws2.cell(row = j, column = 1).value
        d[chat1] += 1
        
    for k in range(2,20):
        ws1.cell(row = k, column = 2).value = d[ws1.cell(row = k, column = 1).value]
    
    wb.save('test.xlsx')
        
    # return lst_student_chat
    print(d)
    
get_student_chat()